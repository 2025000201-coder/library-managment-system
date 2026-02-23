from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from circulation.models import IssuedBook, Fine
from books.models import Book
from accounts.models import User
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from datetime import date


def admin_or_librarian_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.shortcuts import redirect
            return redirect('accounts:login')
        if not (request.user.is_admin_user or request.user.is_librarian_user):
            from django.contrib import messages
            messages.error(request, "Access denied.")
            return redirect('dashboard:home')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
def reports_home(request):
    total_books = Book.objects.count()
    total_students = User.objects.filter(role='student').count()
    total_issued = IssuedBook.objects.filter(status__in=['issued', 'overdue']).count()
    total_overdue = IssuedBook.objects.filter(status='overdue').count()
    total_fines = Fine.objects.filter(status='unpaid').count()

    context = {
        'total_books': total_books,
        'total_students': total_students,
        'total_issued': total_issued,
        'total_overdue': total_overdue,
        'total_fines': total_fines,
    }
    return render(request, 'reports/reports_home.html', context)


# ─────────────────────────────────────────
# EXCEL EXPORTS
# ─────────────────────────────────────────

def style_header_row(ws, row_num, num_cols, fill_color="1a56db"):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align

def add_title(ws, title, num_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="1a56db")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    date_cell = ws.cell(row=2, column=1, value=f"Generated on: {date.today().strftime('%d %B %Y')}")
    date_cell.font = Font(italic=True, color="666666")
    date_cell.alignment = Alignment(horizontal="center")


@admin_or_librarian_required
def export_issued_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Issued Books"

    add_title(ws, "Library Management System — Issued Books Report", 8)

    headers = ["#", "Student Name", "Book Title", "Book ID", "Issue Date", "Due Date", "Return Date", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    status_filter = request.GET.get('status', '')
    queryset = IssuedBook.objects.select_related('student', 'book').all()
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    alt_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    for i, item in enumerate(queryset, 1):
        row = i + 3
        data = [
            i,
            item.student.get_full_name(),
            item.book.title,
            item.book.book_id,
            item.issue_date.strftime('%d-%m-%Y') if item.issue_date else '',
            item.due_date.strftime('%d-%m-%Y') if item.due_date else '',
            item.return_date.strftime('%d-%m-%Y') if item.return_date else 'Not Returned',
            item.get_status_display(),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = alt_fill

    col_widths = [5, 22, 30, 12, 14, 14, 14, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="issued_books_report.xlsx"'
    wb.save(response)
    return response


@admin_or_librarian_required
def export_fines_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fines Report"

    add_title(ws, "Library Management System — Fines Report", 7)

    headers = ["#", "Student Name", "Book Title", "Overdue Days", "Fine/Day (₹)", "Total Fine (₹)", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers), fill_color="c0392b")

    fines = Fine.objects.select_related('student', 'issued_book__book').all()
    alt_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

    for i, fine in enumerate(fines, 1):
        row = i + 3
        data = [
            i,
            fine.student.get_full_name(),
            fine.issued_book.book.title,
            fine.overdue_days,
            float(fine.fine_per_day),
            float(fine.amount),
            fine.get_status_display(),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = alt_fill

    col_widths = [5, 22, 30, 14, 14, 16, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="fines_report.xlsx"'
    wb.save(response)
    return response


@admin_or_librarian_required
def export_books_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books"

    add_title(ws, "Library Management System — Books Catalog", 8)

    headers = ["#", "Book ID", "Title", "Author", "Category", "Total Copies", "Available", "Rack No."]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers), fill_color="1e8449")

    books = Book.objects.select_related('category').all()
    alt_fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")

    for i, book in enumerate(books, 1):
        row = i + 3
        data = [
            i,
            book.book_id,
            book.title,
            book.author,
            book.category.name if book.category else '—',
            book.total_copies,
            book.available_copies,
            book.rack_number or '—',
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = alt_fill

    col_widths = [5, 12, 35, 22, 16, 14, 12, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="books_catalog.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────
# PDF EXPORTS
# ─────────────────────────────────────────

def get_pdf_styles():
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                  fontSize=18, textColor=colors.HexColor('#1a56db'),
                                  spaceAfter=4, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'],
                                     fontSize=10, textColor=colors.grey,
                                     spaceAfter=20, alignment=TA_CENTER)
    return title_style, subtitle_style


@admin_or_librarian_required
def export_issued_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                             rightMargin=1.5*cm, leftMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)

    title_style, subtitle_style = get_pdf_styles()
    elements = []
    elements.append(Paragraph("Library Management System", title_style))
    elements.append(Paragraph(f"Issued Books Report — {date.today().strftime('%d %B %Y')}", subtitle_style))

    headers = ["#", "Student", "Book Title", "Book ID", "Issue Date", "Due Date", "Return Date", "Status"]
    data = [headers]

    queryset = IssuedBook.objects.select_related('student', 'book').all()
    for i, item in enumerate(queryset, 1):
        data.append([
            str(i),
            item.student.get_full_name(),
            item.book.title[:35],
            item.book.book_id,
            item.issue_date.strftime('%d-%m-%Y') if item.issue_date else '',
            item.due_date.strftime('%d-%m-%Y') if item.due_date else '',
            item.return_date.strftime('%d-%m-%Y') if item.return_date else 'Pending',
            item.get_status_display(),
        ])

    table = Table(data, colWidths=[1*cm, 4.5*cm, 6*cm, 2.5*cm, 2.8*cm, 2.8*cm, 2.8*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a56db')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#EEF2FF')]),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWHEIGHT', (0,0), (-1,-1), 22),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="issued_books_report.pdf"'
    return response


@admin_or_librarian_required
def export_fines_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             rightMargin=1.5*cm, leftMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)

    title_style, subtitle_style = get_pdf_styles()
    elements = []
    elements.append(Paragraph("Library Management System", title_style))
    elements.append(Paragraph(f"Fines Report — {date.today().strftime('%d %B %Y')}", subtitle_style))

    headers = ["#", "Student", "Book", "Overdue Days", "Fine/Day", "Total Fine", "Status"]
    data = [headers]

    fines = Fine.objects.select_related('student', 'issued_book__book').all()
    total_unpaid = sum(f.amount for f in fines if f.status == 'unpaid')

    for i, fine in enumerate(fines, 1):
        data.append([
            str(i),
            fine.student.get_full_name(),
            fine.issued_book.book.title[:30],
            str(fine.overdue_days),
            f"Rs.{fine.fine_per_day}",
            f"Rs.{fine.amount}",
            fine.get_status_display(),
        ])

    data.append(['', '', '', '', '', f"Total Unpaid: Rs.{total_unpaid}", ''])

    table = Table(data, colWidths=[1*cm, 5*cm, 5.5*cm, 3*cm, 2.5*cm, 3.5*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#c0392b')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#FDEDEC')]),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWHEIGHT', (0,0), (-1,-1), 22),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#FDEDEC')),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="fines_report.pdf"'
    return response


@admin_or_librarian_required
def export_overdue_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                             rightMargin=1.5*cm, leftMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)

    title_style, subtitle_style = get_pdf_styles()
    elements = []
    elements.append(Paragraph("Library Management System", title_style))
    elements.append(Paragraph(f"Overdue Books Report — {date.today().strftime('%d %B %Y')}", subtitle_style))

    headers = ["#", "Student", "Contact", "Book Title", "Book ID", "Due Date", "Days Overdue", "Est. Fine"]
    data = [headers]

    overdue = IssuedBook.objects.filter(status='overdue').select_related('student', 'book')
    for i, item in enumerate(overdue, 1):
        data.append([
            str(i),
            item.student.get_full_name(),
            item.student.phone or '—',
            item.book.title[:35],
            item.book.book_id,
            item.due_date.strftime('%d-%m-%Y'),
            str(item.overdue_days),
            f"Rs.{item.overdue_days * 2}",
        ])

    table = Table(data, colWidths=[1*cm, 4.5*cm, 3*cm, 6*cm, 2.5*cm, 3*cm, 3*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e67e22')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#FEF9E7')]),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWHEIGHT', (0,0), (-1,-1), 22),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="overdue_report.pdf"'
    return response